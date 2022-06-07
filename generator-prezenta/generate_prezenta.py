from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from generate_libere import Libere
from datetime import date
import calendar, math, sys
_month = {"ianuarie": 1, "februarie": 2, "martie": 3, "aprilie": 4, "mai": 5, "iunie": 6, "iulie": 7, "august": 8, "septembrie": 9, "octombrie": 10, "noiembrie": 11, "decembrie": 12}

def get_work_days(month, year):
    max = calendar.monthrange(year, month)[1]
    lista = []
    for day in range(1, max+1):
        weekday = calendar.weekday(year, month, day)
        if weekday < calendar.SATURDAY:
            if day < 10:
                day = "0" + str(day)
            else:
                day = str(day)
            lista.append(day)
    return lista

class Excel:
    def __init__(self, output_folder=""):
        self.wb = Workbook()
        sh = self.wb.active
        sh.title = "Prezenta"
        self.sh = self.wb[sh.title]
        self.month = Excel.value_to_key(date.today().month)
        self.year = date.today().year
        self.default = Excel.default() #dictionary
        self.output_folder = output_folder

    def set_default(self, font_name=None, font_size=None, font_bold=None, border=None):
        if font_name is not None:
            self.default["font_name"] = font_name
        if font_size is not None:
            self.default["font_size"] = font_size
        if font_bold is not None:
            self.default["font_bold"] = font_bold
        if border is not None:
            self.default["boder"] = border

    def get_font_name(self):
        return self.default["font_name"]

    def get_font_size(self):
        return self.default["font_size"]

    def get_font_bold(self):
        return self.default["font_bold"]

    @staticmethod
    def transformSelection(selection): # tuple (row, column) -> col+row
        if isinstance(selection, str):
            return selection # do nothing

        row = selection[0]
        column = selection[1]
        if column > 16384 or column < 1 or row > 1048576 or row < 1:
            raise Exception("Value exceed excel board.")
        letter = Excel.get_column_letter(column)

        return "".join(letter) + str(row)

    @staticmethod
    def get_column_letter(column):
        if isinstance(column, int): # 1
            col = [] # A, B, AA, AB, BA, BC
            alfa = "ABCDEFGHIJKLMNOPQRSTUVWXYZ" # 26
            base = len(alfa)
            n = int(math.log(column, base)) # 2 1 0
            while column > 0:
                product  = 1 ; n1 = n
                while n1 > 0: # base ** n
                    product *= base
                    n1 -= 1
                value = int(column/product)
                col.append(alfa[value-1])
                column = column % product
                n -= 1
            return "".join(col)
        elif isinstance(column, str): # A12
            result = "".join([i for i in column if i is not i.isdigit()])
            return result #A

    def add_header_image(self, selection="B1", path_image="src/allora.PNG"):
        if isinstance(selection, tuple) or isinstance(selection, list):
            select = Excel.transformSelection(selection)
        elif not(isinstance(selection, str)):
            raise Exception("Not String or tuple / list")
        else:
            select = selection
        img = Image(path_image)
        self.sh.add_image(img, select)

    def add_value(self, selection, value=None):
        if isinstance(selection, str):
            self.sh[selection] = value
            self.sh[selection].font = Font(name=self.default["font_name"], size=self.default["font_size"], bold=self.default["font_bold"])
        elif isinstance(selection, tuple) or isinstance(selection, list):
            row = selection[0]
            column = selection[1]
            self.sh.cell(row=row, column=column, value=value)
            self.sh.cell(row=row, column=column).font = Font(name=self.default["font_name"], size=self.default["font_size"], bold=self.default["font_bold"])
        else:
            raise Exception("valueType")

    @staticmethod
    def default():
        return {"font_name": "Times New Roman",
                "font_size": "12",
                "font_bold": False,
                "border": {
                    "left": ["thin", 'FF000000'],
                    "right": ["thin", 'FF000000'],
                    "top": ["thin", 'FF000000'],
                    "bottom": ["thin", 'FF000000']}
                }

    def font_style(self, selection, name=None, bold=False, size=None, wrap=False):
        if isinstance(selection, str):
            select = self.sh[selection]
        elif isinstance(selection, tuple) or isinstance(selection, list):
            select = self.sh.cell(row=selection[0], column=selection[1])
        else:
            select = None
            raise Exception("No selecting - font_style")

        select.font = Font(name=self.default["font_name"] if name is None else name, size=size if size is not None else self.default["font_size"], bold=True if bold else self.default["font_bold"])

    def font_align(self, selection, wrap = False, horizontal=None, vertical =None):
        if isinstance(selection, str):
            select = self.sh[selection]
        elif isinstance(selection, tuple) or isinstance(selection, list):
            select = self.sh.cell(row=selection[0], column=selection[1])
        else:
            select = None
            raise Exception("No selecting - font_align")

        select.alignment = Alignment(wrap_text=True if wrap else False, horizontal=horizontal if horizontal is not None else 'left', vertical=vertical if vertical is not None else 'bottom')

    def border_style(self, selection, listBorder=[]): # ma mai gandesc la border default
        if isinstance(selection, str): # [[thin, ffff], [thin, ffff], [], []] n-e-s-w
            select = self.sh[selection]
        elif isinstance(selection, tuple) or isinstance(selection, list):
            select = self.sh.cell(row=selection[0], column=selection[1])
        else:
            select = None
            raise Exception("No selecting - border_style")
        sideLeft = Side(style=self.default["border"]["left"][0], color=self.default["border"]["left"][1]) if len(listBorder) < 4 else Side(style=listBorder[3][0], color=listBorder[3][1])
        sideRight = Side(style=self.default["border"]["right"][0], color=self.default["border"]["right"][1]) if len(listBorder) < 4 else Side(style=listBorder[1][0], color=listBorder[1][1])
        sideTop = Side(style=self.default["border"]["top"][0], color=self.default["border"]["top"][1]) if len(listBorder) < 4 else Side(style=listBorder[0][0], color=listBorder[0][1])
        sideBottom = Side(style=self.default["border"]["bottom"][0], color=self.default["border"]["bottom"][1]) if len(listBorder) < 4 else Side(style=listBorder[2][0], color=listBorder[2][1])
        select.border = Border(left=sideLeft, right=sideRight, top=sideTop, bottom=sideBottom)

    def modify_column(self, column, value):
        if isinstance(column, str):
            self.sh.column_dimensions[column].width = value
        elif isinstance(column, int):
            self.sh.column_dimensions[Excel.get_column_letter(column)].width = value

    def modify_row(self, row, value):
        self.sh.row_dimensions[row].height = value

    def merge_cells(self, selection=None):
        if isinstance(selection, str):
            self.sh.merge_cells(range_string=selection)
        elif isinstance(selection, list) or isinstance(selection, tuple):
            if isinstance(selection[0], tuple) and isinstance(selection[1], tuple): # supose selecton = [ [1,2], [3,4]]
                self.sh.merge_cells(range_string=None, start_row=selection[0][0], start_column=selection[0][1], end_row=selection[1][0], end_column=selection[1][1])
            else: # [1, 2, 3, 4]
                self.sh.merge_cells(range_string=None, start_row=selection[0], start_column=selection[1], end_row=selection[2], end_column=selection[3])
        else:
            raise Exception("merge cells - Wrong selection: {}".format(selection))

    @staticmethod
    def value_to_key(value): # Aprilie -> 4
        listKeys = list(_month.keys())
        listValues = list(_month.values())
        return listKeys[listValues.index(value)]

    def save(self):
        self.wb.save("{}Prezenta-{}.xlsx".format(self.output_folder, self.month))

if __name__ == '__main__':
    current_year = date.today().year
    current_month = date.today().month

    current_month_str = str(current_month) if current_month > 9 else "0" + str(current_month)
    days = get_work_days(current_month, current_year)

    # getting free legal days
    free_days = []

    settings = Libere.loadJson("settings.json")
    persons = settings["persons"]
    dim_columns = settings["dim_columns"]
    introductionData = list(settings["company_data"].values())
    ORA_INCEPUT, ORA_SFARSIT, PAUZA = list(settings["schedule"].values())
    INPUT_FOLDER = settings["input_folder"]
    logo_company = settings["company_data"]["logo_img_path"]

    try:
        with open(INPUT_FOLDER + "libere_{}.txt".format(current_year), "r") as file:
            for item in file.readlines():
                day, month, year = list(map(int, item.strip().split("/")))
                if month  == current_month:
                    free_days.append(day)
    except FileNotFoundError:
        print("Generate free_days first.")
        sys.exit("Generate free_days first.")



    wb = Excel(settings["output_folder"])

    #adaugat imagine
    wb.add_header_image(path_image=logo_company)

    # adaugat header text
    for i in range(4, 12):
        wb.add_value([i, 2], introductionData[i-4])
        if i == 8:
            wb.font_align([i, 2], wrap=True, vertical="center")
            wb.modify_row(i, 50)
    wb.add_value("D12", "CONDICA PREZENTA")
    wb.merge_cells("D12:G12")
    wb.font_align("D12", horizontal="left")

    wb.font_style("B10", bold=True, size="10")
    wb.font_style("B11", bold=True)
    wb.font_style("D12", bold=True, size="14")

    #completat tabel
    curentRow = 15
    wb.set_default(font_name="Calibri", font_size="11")
    for section in range(len(days)):
        for row in range(1, 2+len(persons) +1): #6
            for col in range(1, 8+1):
                selection = [curentRow, col]
                if row == 1:
                    if col == 2:
                        wb.add_value(selection, "ZIUA:{}".format(days[section]))
                        wb.font_style(selection, name=wb.get_font_name(), size=wb.get_font_size(), bold=True)
                    elif col == 4:
                        wb.add_value(selection, "LUNA:{}".format(current_month_str))
                        wb.font_style(selection, name=wb.get_font_name(), size=wb.get_font_size(), bold=True)
                    elif col == 6:
                        wb.add_value(selection, "ANUL:{}".format(current_year))
                        wb.font_style(selection, name=wb.get_font_name(), size=wb.get_font_size(), bold=True)
                elif row == 2:
                    if col == 1:
                        wb.add_value(selection, "Nr.")
                        wb.font_align(selection, vertical="center")
                    elif col == 2:
                        wb.add_value(selection, "Nume si prenume")
                        wb.font_align(selection, vertical="center")
                    elif col == 3:
                        wb.add_value(selection, "Semnat. Venire")
                        wb.font_align(selection, wrap=True)
                    elif col == 4 or col == 6:
                        wb.add_value(selection, "Ora")
                        wb.font_align(selection, horizontal="center", vertical="center")
                    elif col == 5:
                        wb.add_value(selection, "Semnat plecare")
                        wb.font_align(selection, wrap=True)
                    elif col == 7:
                        wb.add_value(selection, "Pauza de masa")
                        wb.font_align(selection, wrap=True)
                    else: #8
                        wb.add_value(selection, "Observatii")
                        wb.font_align(selection, vertical="center")
                    wb.modify_row(curentRow, 30)
                else: # > 3..6
                    if col == 1:
                        wb.add_value(selection, row-2)
                    elif col == 2:
                        wb.add_value(selection, persons[row-3])
                    elif col == 4:
                        wb.add_value(selection, ORA_INCEPUT)
                        wb.font_align(selection, horizontal="center")
                    elif col == 6:
                        wb.add_value(selection, ORA_SFARSIT)
                        wb.font_align(selection, horizontal="center")
                    elif col == 7:
                        wb.add_value(selection, PAUZA)
                    elif col == 8: # obs data
                        if row == 3: # first data value
                            if free_days.count(int(days[section])) > 0:
                                wb.add_value(selection, "Zi libera legala")
                                wb.merge_cells(selection=((curentRow, 8), (curentRow + len(persons) -1 , 8)))
                                wb.font_style(selection, bold=True)
                                wb.font_align(selection, wrap=True, horizontal="center", vertical="center")

                wb.border_style(selection)
            curentRow += 1

    # setat dimensiune coloane
    for col in range(1, 8+1): # -> h
        wb.modify_column(col, dim_columns[col-1])

    wb.save()
